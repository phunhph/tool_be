import os
import shutil
import zipfile
import asyncio
import logging
import tempfile
import json
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List, Dict, Any, Callable, Awaitable, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

from fastapi import UploadFile, HTTPException
from fastapi.responses import StreamingResponse, FileResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
from sqlalchemy.orm import Session, joinedload

from app.models import Report, Exam
from app.schemas.report import ReportStatus, ReportCreate, ReportUpdate
from app.schemas.base_schemas import (
    CreateResponse, DeleteResponse, DetailResponse, ListResponse, UpdateResponse
)
from app.services.gemini_service import GeminiService, PLAGIARISM_THRESHOLD
from app.tasks.report_tasks import process_uploaded_archive
from app.core.config import settings
from app.core.google_drive import get_google_drive_manager
# Logging setup
logger = logging.getLogger(__name__)

STORAGE_ROOT = Path(settings.REPORT_STORAGE_ROOT)
STORAGE_ROOT.mkdir(parents=True, exist_ok=True)
MAX_WORKERS = 5  # số luồng xử lý song song


# ------------------- Helper chung -------------------
def raise_error(status: int, message: str):
    """Raise HTTPException với message rõ ràng"""
    raise HTTPException(status_code=status, detail=message)


# ------------------- Service chính -------------------
class ReportService:

    # ------------------- Mapper -------------------
    @staticmethod
    def map_to_schema(report: Report) -> Dict[str, Any]:
        return {
            "id": report.id,
            "name": report.name,
            "student_code": report.student_code,
            "major": report.major,
            "position": report.position,
            "advantage": report.strengths,
            "disadvantage": report.weaknesses,
            "suggestion": report.proposal,
            "note": report.note,
            "attitude_point": report.attitude_score,
            "work_point": report.work_score,
            "status": report.status,
            "exam_id": report.exam_id,
            "created_at": report.created_at,
            "error": report.processing_error,
            "files": [
                {
                    "id": f.id,
                    "name_file": f.name_file,
                    "path_storage": f.path_storage,
                    "created_at": f.created_at,
                }
                for f in getattr(report, "files", [])
            ],
        }

    # ------------------- CRUD -------------------
    @staticmethod
    def get_list(db: Session, page: int, page_size: int, exam_id: Optional[int] = None):
        query = db.query(Report).options(
            joinedload(Report.files),  # Eager load files
            joinedload(Report.exam)    # Eager load exam nếu cần
        ).order_by(Report.created_at.desc())

        if exam_id is not None:
            query = query.filter(Report.exam_id == exam_id)

        total = query.count()
        reports = query.offset((page - 1) * page_size).limit(page_size).all()

        return ListResponse(
            data=[ReportService.map_to_schema(r) for r in reports],
            total=total,
            pageSize=page_size,
            pageIndex=page,
        )

    @staticmethod
    def get_detail(db: Session, report_id: int):
        report = db.query(Report).options(joinedload(Report.files)).filter(Report.id == report_id).first()
        if not report:
            raise_error(404, "Report không tồn tại")
        return DetailResponse(status=True, data=ReportService.map_to_schema(report))

    @staticmethod
    def create(db: Session, payload: ReportCreate, username: str):
        data = payload.dict(exclude_unset=True)
        new_report = Report(
            name=data.get("name"),
            student_code=data.get("student_code"),
            major=data.get("major"),
            position=data.get("position"),
            strengths=data.get("advantage"),
            weaknesses=data.get("disadvantage"),
            proposal=data.get("suggestion"),
            attitude_score=data.get("attitude_point"),
            work_score=data.get("work_point"),
            note=data.get("note"),
            status=data.get("status") or ReportStatus.pending,
            exam_id=data.get("exam_id"),
            created_at=datetime.utcnow(),
            created_by=username,
            processing_error=data.get("error"),
        )
        db.add(new_report)
        db.commit()
        db.refresh(new_report)

        return CreateResponse(message="Tạo báo cáo thành công", status=True, objectId=new_report.id)

    @staticmethod
    def update(db: Session, report_id: int, payload: ReportUpdate):
        report = db.query(Report).filter(Report.id == report_id).first()
        if not report:
            raise_error(404, "Report không tồn tại")

        for key, value in payload.dict(exclude_unset=True).items():
            if key == "error":
                report.processing_error = value
            else:
                setattr(report, key, value)

        db.commit()
        db.refresh(report)

        return UpdateResponse(
            message="Cập nhật báo cáo thành công",
            status=True,
            data=ReportService.map_to_schema(report),
        )

    @staticmethod
    def delete(db: Session, report_id: int):
        report = db.query(Report).filter(Report.id == report_id).first()
        if not report:
            raise_error(404, "Report không tồn tại")

        db.delete(report)
        db.commit()

        return DeleteResponse(message="Xóa báo cáo thành công", status=True, examId=report.id)

    # ------------------- Worker xử lý file đơn -------------------
    @staticmethod
    def _process_single_file_task_worker(file_data: Dict[str, Any], folder_path: str) -> Dict[str, Any]:
        file_name = file_data["filename"]
        file_path = os.path.join(folder_path, file_name)

        try:
            with open(file_path, "rb") as f:
                pdf_bytes = f.read()

            info = GeminiService.extract_info_from_pdf(pdf_bytes)

            return {
                "original_filename": file_name,
                "name": info.get("Họ và tên", file_name),
                "student_code": info.get("MSSV", "UNKNOWN"),
                "raw_content": info.get("Nội dung báo cáo thô", ""),
                "status": "SUCCESS" if info.get("MSSV") else "MISSING_MSSV",
                **info,
            }
        except Exception as e:
            return {"original_filename": file_name, "status": "ERROR", "error_detail": str(e)}

    # ------------------- Hàm xử lý đa luồng chính -------------------
    @staticmethod
    async def run_full_processing(task, exam_id: str, folder_path: str, file_metadata: List[Dict[str, Any]], update_progress_func):
        """
        Hàm xử lý chính chạy trong Celery Worker.
        Gồm các bước:
             Duyệt từng file PDF
             Gọi Gemini để trích xuất thông tin
             Kiểm tra đạo văn
             Cập nhật tiến độ lên Redis + WebSocket
             Nén kết quả thành ZIP
        """
        total = len(file_metadata)
        processed = 0
        results = []
        plagiarism_results = []

        # Tạo folder tạm cho output
        temp_output_dir = tempfile.mkdtemp(prefix=f"reports_{exam_id}_")

        for file_info in file_metadata:
            try:
                filename = file_info["filename"]
                file_path = os.path.join(folder_path, filename)

                # --- Đọc file ---
                with open(file_path, "rb") as f:
                    pdf_bytes = f.read()

                # --- Trích xuất dữ liệu từ Gemini ---
                extracted = GeminiService.extract_info_from_pdf(pdf_bytes)
                extracted["filename"] = filename
                results.append(extracted)

                processed += 1

                # --- Gửi tiến độ ---
                await update_progress_func(task, processed, total, exam_id)

            except Exception as e:
                print(f"[ERROR] Lỗi xử lý file {filename}: {e}")
                continue

        # --- Kiểm tra đạo văn (nếu có >1 file) ---
        if len(results) > 1:
            for i in range(len(results)):
                for j in range(i + 1, len(results)):
                    sim = GeminiService.check_plagiarism_similarity(
                        results[i].get("Nội dung báo cáo thô", ""),
                        results[j].get("Nội dung báo cáo thô", "")
                    )
                    if sim >= 0.75:
                        plagiarism_results.append({
                            "file_a": results[i]["filename"],
                            "file_b": results[j]["filename"],
                            "similarity": round(sim, 3)
                        })

        # --- Ghi kết quả ra JSON ---
        output_json_path = os.path.join(temp_output_dir, "summary.json")
        with open(output_json_path, "w", encoding="utf-8") as f:
            json.dump({
                "exam_id": exam_id,
                "results": results,
                "plagiarism_results": plagiarism_results
            }, f, ensure_ascii=False, indent=2)

        # --- Nén thành ZIP ---
        zip_output_path = os.path.join(temp_output_dir, f"exam_{exam_id}_result.zip")
        with zipfile.ZipFile(zip_output_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(temp_output_dir):
                for file in files:
                    zipf.write(os.path.join(root, file), file)

        # --- Gửi tiến độ hoàn tất ---
        await update_progress_func(task, total, total, exam_id)

        # --- Trả kết quả cuối cùng cho Celery backend ---
        return {
            "plagiarism_results": plagiarism_results,
            "zip_file": zip_output_path
        }

    # ------------------- Nhận file ZIP và khởi tạo task -------------------
    @staticmethod
    def process_zip_upload(db: Session, exam_id: int, zip_bytes: bytes, zip_filename: str, username: str):
        """
        Xử lý upload file ZIP chứa các file PDF báo cáo.
        - Upload trực tiếp lên Google Drive (nếu enabled)
        - Fallback: lưu local nếu Google Drive không cấu hình
        - Khởi tạo Celery task để xử lý bất đồng bộ
        """
        try:
            # Kiểm tra exam tồn tại
            exam = db.query(Exam).filter(Exam.id == exam_id).first()
            if not exam:
                raise_error(404, "Kỳ thi không tồn tại")

            # Tạo timestamp và folder name
            timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
            folder_name = f"report_{exam.code}_{timestamp}"
            
            # Kiểm tra có nên dùng Google Drive không
            use_gdrive = settings.GOOGLE_DRIVE_ENABLED and settings.GOOGLE_DRIVE_ROOT_FOLDER_ID
            gd_manager = get_google_drive_manager() if use_gdrive else None
            gdrive_folder_id = None
            folder_path = None
            gdrive_upload_failed = False  # Flag: nếu upload Google Drive fail, không thử lại
            
            # Tạo folder local ngay (không chờ test Google Drive)
            folder_path = STORAGE_ROOT / folder_name
            try:
                folder_path.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                logger.error(f"Lỗi tạo folder local: {e}")
                raise_error(500, f"Không thể tạo folder lưu file: {e}")
            
            # Nếu Google Drive enabled và configured, thử tạo folder trên GDrive
            if use_gdrive and gd_manager and gd_manager.is_configured:
                gdrive_folder_id = gd_manager.create_folder(
                    folder_name,
                    parent_id=settings.GOOGLE_DRIVE_ROOT_FOLDER_ID
                )
                if not gdrive_folder_id:
                    logger.warning("Không thể tạo folder trên Google Drive, sẽ dùng local storage")
                    use_gdrive = False
                    gdrive_upload_failed = True

            file_metadata = []
            max_file_size = 100 * 1024 * 1024  # 100MB per file

            try:
                # Giải nén và kiểm tra file ZIP
                with zipfile.ZipFile(BytesIO(zip_bytes), "r") as zip_ref:
                    # Kiểm tra ZIP có hợp lệ không
                    bad_file = zip_ref.testzip()
                    if bad_file:
                        raise zipfile.BadZipFile(f"File ZIP bị lỗi tại: {bad_file}")

                    # Lọc chỉ lấy file PDF
                    member_names = [
                        name for name in zip_ref.namelist() 
                        if name.lower().endswith(".pdf") and not name.endswith("/")
                    ]

                    if not member_names:
                        if use_gdrive and gdrive_folder_id:
                            # Xoá folder trên Google Drive nếu không có file hợp lệ
                            logger.warning(f"Xoá folder Google Drive không có file: {gdrive_folder_id}")
                        raise_error(400, "File ZIP không chứa file PDF hợp lệ.")

                    # Giới hạn số lượng file
                    if len(member_names) > 100:
                        raise_error(400, "File ZIP chứa quá nhiều file PDF (tối đa 100 file).")

                    # Giải nén từng file
                    for name in member_names:
                        try:
                            # Kiểm tra kích thước file
                            file_info = zip_ref.getinfo(name)
                            if file_info.file_size > max_file_size:
                                logger.warning(f"File {name} quá lớn ({file_info.file_size} bytes), bỏ qua")
                                continue

                            content = zip_ref.read(name)
                            base_name = os.path.basename(name)
                            
                            # Làm sạch tên file (loại bỏ ký tự đặc biệt)
                            base_name = "".join(c for c in base_name if c.isalnum() or c in "._- ()[]")
                            if not base_name:
                                base_name = f"file_{datetime.utcnow().timestamp()}.pdf"
                            
                            # Đảm bảo có extension .pdf
                            if not base_name.lower().endswith(".pdf"):
                                base_name += ".pdf"

                            # Upload lên Google Drive nếu enabled và chưa fail
                            if use_gdrive and gd_manager and gdrive_folder_id and not gdrive_upload_failed:
                                file_id = gd_manager.upload_file_bytes(
                                    content,
                                    base_name,
                                    parent_id=gdrive_folder_id
                                )
                                if file_id:
                                    file_metadata.append({
                                        "filename": base_name,
                                        "path": f"gdrive://{file_id}",
                                        "gdrive_file_id": file_id,
                                        "size": len(content),
                                        "storage": "google_drive"
                                    })
                                    logger.info(f"[GDRIVE] Uploaded: {base_name} -> {file_id}")
                                else:
                                    # Fallback: Google Drive upload thất bại lần đầu, sẽ lưu local cho file này và file sau
                                    logger.warning(f"[FALLBACK] Google Drive upload failed for {base_name}, will save to local storage from now on")
                                    gdrive_upload_failed = True
                                    
                                    temp_path = folder_path / base_name
                                    counter = 1
                                    while temp_path.exists():
                                        name_part = os.path.splitext(base_name)[0]
                                        ext_part = os.path.splitext(base_name)[1]
                                        base_name = f"{name_part}_{counter}{ext_part}"
                                        temp_path = folder_path / base_name
                                        counter += 1
                                    
                                    with open(temp_path, "wb") as f:
                                        f.write(content)
                                    
                                    file_metadata.append({
                                        "filename": base_name,
                                        "path": str(temp_path.resolve()),
                                        "size": len(content),
                                        "storage": "local"
                                    })
                                    logger.info(f"[LOCAL] Saved (fallback): {base_name} at {temp_path}")
                            else:
                                # Lưu local
                                temp_path = folder_path / base_name

                                # Tránh ghi đè file trùng tên
                                counter = 1
                                while temp_path.exists():
                                    name_part = os.path.splitext(base_name)[0]
                                    ext_part = os.path.splitext(base_name)[1]
                                    base_name = f"{name_part}_{counter}{ext_part}"
                                    temp_path = folder_path / base_name
                                    counter += 1
                                    
                                    if counter > 1000:
                                        base_name = f"{datetime.utcnow().timestamp()}_{base_name}"
                                        temp_path = folder_path / base_name
                                        break

                                # Lưu file
                                with open(temp_path, "wb") as f:
                                    f.write(content)

                                file_metadata.append({
                                    "filename": base_name,
                                    "path": str(temp_path.resolve()),
                                    "size": len(content),
                                    "storage": "local"
                                })
                                
                                logger.info(f"[LOCAL] Saved: {base_name} at {temp_path}")

                        except Exception as e:
                            logger.error(f"Lỗi khi giải nén file {name}: {e}")
                            continue

                if not file_metadata:
                    raise_error(400, "Không có file PDF hợp lệ nào trong ZIP.")

            except zipfile.BadZipFile as e:
                logger.error(f"File ZIP không hợp lệ: {e}")
                raise_error(400, f"File ZIP không hợp lệ: {e}")
            except Exception as e:
                logger.error(f"Lỗi khi giải nén file ZIP: {e}", exc_info=True)
                raise_error(500, f"Lỗi khi giải nén file ZIP: {e}")

            # Gọi Celery task để xử lý bất đồng bộ
            try:
                clean_metadata = [
                    {
                        "filename": str(item.get("filename")),
                        "path": str(item.get("path")),
                        "storage": item.get("storage", "local")
                    }
                    for item in file_metadata
                ]

                clean_folder_path = str(folder_path) if folder_path else f"gdrive://{gdrive_folder_id}"

                task = process_uploaded_archive.delay(
                    str(exam_id),
                    clean_folder_path,
                    clean_metadata,
                    str(username),
                )
                logger.info(
                    f"[TASK] Khởi tạo task {task.id} để xử lý {len(clean_metadata)} file PDF từ {clean_folder_path}"
                )
            except Exception as e:
                logger.error(f"Lỗi khởi tạo Celery task: {e}", exc_info=True)
                raise_error(500, f"Không thể khởi tạo task xử lý: {e}")

            return {
                "success": True,
                "status": True,
                "objectId": exam_id,
                "message": f"Đã nhận {len(file_metadata)} file PDF. Hệ thống đang xử lý ngầm.",
                "data": {
                    "task_id": task.id,
                    "websocket_url": f"/ws/status/{task.id}",
                    "api_url": f"/tasks/status/{task.id}",
                    "file_count": len(file_metadata),
                    "storage": "google_drive" if use_gdrive else "local"
                },
            }

        except HTTPException:
            raise
        except Exception as e:
            logger.error(f"Lỗi không mong đợi trong process_zip_upload: {e}", exc_info=True)
            raise_error(500, f"Lỗi xử lý upload: {e}")
    
    @staticmethod
    def export_by_exam(db, exam_id: int):
        # Lấy dữ liệu báo cáo theo kỳ thi
        reports = db.query(Report).filter(
            Report.exam_id == exam_id
        ).order_by(Report.created_at.desc()).all()

        if not reports:
            return {"status": False, "message": "Không có dữ liệu để xuất Excel."}

        # === TẠO FILE EXCEL ===
        wb = Workbook()
        ws = wb.active
        ws.title = f"Exam_{exam_id}"

        # Header
        headers = [
            "STT", "Họ và tên", "MSSV", "Ngành", "Công ty thực tập", "Vị trí thực tập",
            "Ưu điểm", "Nhược điểm", "Đề xuất", "Điểm thái độ", "Điểm công việc",
            "Ghi chú", "Ngày tạo", "Người tạo"
        ]
        ws.append(headers)

        # Style cho header
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        # Ghi dữ liệu
        for i, r in enumerate(reports, start=1):
            row = [
                i,
                r.name or "",
                r.student_code or "",
                r.major or "",
                r.company or "",
                r.position or "",
                r.strengths or "",
                r.weaknesses or "",
                r.proposal or "",
                r.attitude_score or "",
                r.work_score or "",
                r.note or "",
                r.created_at.strftime("%Y-%m-%d %H:%M:%S") if r.created_at else "",
                r.created_by or "",
            ]
            ws.append(row)

        # Căn giữa và thêm viền
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

        # Auto width
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(length + 2, 50)

        # Ghi ra buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Tên file
        filename = f"Bao_cao_thuc_tap_{exam_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Trả file về client
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )

    @staticmethod
    def download_report_pdf(db: Session, report_id: int) -> FileResponse:
        report = db.query(Report).options(joinedload(Report.files)).filter(Report.id == report_id).first()
        if not report:
            raise_error(404, "Report không tồn tại")

        if not report.files:
            raise_error(404, "Báo cáo không có file đính kèm")

        file_record = report.files[0]
        file_path = Path(file_record.path_storage)

        if not file_path.exists():
            raise_error(404, "File báo cáo không tồn tại trên hệ thống")

        return FileResponse(
            path=str(file_path),
            media_type="application/pdf",
            filename=file_record.name_file or file_path.name
        )